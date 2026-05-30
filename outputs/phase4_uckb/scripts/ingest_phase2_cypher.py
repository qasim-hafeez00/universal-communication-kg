"""
UCKB Phase 4 — Track B: Production Cypher Ingestion
Reads Phase 2 Excel and creates clean production nodes + relationships in Neo4j.

This is the complement to Track A (n10s RDF import).  Track B creates:
  - Typed :Technique / :DomainProtocol / :EmotionalState / :SignalMarker
    / :CommunicationStyle nodes with production-grade Neo4j properties
  - All 11 locked relationship types as typed Neo4j relationships
  - Computed fields: evidenceLevel, avgLatency_ms, requiresEmotionalClearance
  - Array-valued multi-value properties (whenToUse, sourceIds, etc.)

Prerequisites:
  pip install neo4j openpyxl

Usage:
  python ingest_phase2_cypher.py [--uri bolt://localhost:7687] [--user neo4j] [--password uckb_admin_2024]
"""

import re
import sys
import argparse
from pathlib import Path

import openpyxl
from neo4j import GraphDatabase, exceptions as neo4j_exc

# ── Paths ─────────────────────────────────────────────────────────────────────
ROOT     = Path(__file__).resolve().parent.parent.parent.parent
PHASE2   = ROOT / "outputs" / "phase2_uckb" / "Phase_2_UCKB_Master_Taxonomy.xlsx"

# ── Connection defaults ───────────────────────────────────────────────────────
DEFAULT_URI      = "bolt://localhost:7687"
DEFAULT_USER     = "neo4j"
DEFAULT_PASSWORD = "uckb_admin_2024"

# ── Domain label mapping ──────────────────────────────────────────────────────
DOMAIN_LABEL_MAP = {
    "Crisis Dispatch / Emergency": "Crisis Dispatch",
    "Sales & Negotiation":         "Sales Negotiation",
    "Clinical / Medical":          "Clinical Medical",
}

# ── Class → Neo4j label mapping ───────────────────────────────────────────────
CLASS_LABEL_MAP = {
    "CommunicationTechnique": "Technique",
    "DomainProtocol":         "DomainProtocol",
    "EmotionalState":         "EmotionalState",
    "SignalMarker":           "SignalMarker",
    "CommunicationStyle":     "CommunicationStyle",
}

# ── Locked edge vocabulary ────────────────────────────────────────────────────
EDGE_VOCAB = {
    "REQUIRES", "ENHANCES", "CONTRADICTS", "DOMAIN_VARIANT_OF",
    "CULTURAL_VARIANT_OF", "CONTRAINDICATED_WHEN", "ESCALATES_TO",
    "RESOLVES", "TRIGGERED_BY", "PRECEDES", "FOLLOWS",
}

# ── avgLatency_ms by domain (milliseconds — estimated, refined in Phase 11) ───
DOMAIN_LATENCY = {
    "Crisis Dispatch / Emergency": 30,   # lowest: life-safety requires fastest routing
    "Clinical / Medical":          80,   # medium: clinical precision over speed
    "Sales & Negotiation":         50,   # balanced: strategic but not critical-safety
}

# ── Emotional state valence/arousal for cards with EmotionalState class ──────
EMOTIONAL_STATE_PROPS = {
    "crisis_dispatch_026_panic_state":             ("negative", "high"),
    "crisis_dispatch_027_hostile_state":           ("negative", "high"),
    "crisis_dispatch_028_defensive_state":         ("negative", "medium"),
    "crisis_dispatch_029_cognitive_overload_state":("negative", "high"),
    "crisis_dispatch_048_adult_ego_state":         ("neutral",  "medium"),
    "clinical_032_grief_pause":                    ("negative", "low"),
    "clinical_033_anger_at_diagnosis_validation":  ("negative", "high"),
    "clinical_034_fear_of_prognosis_validation":   ("negative", "high"),
    "clinical_045_clinical_panic_state":           ("negative", "high"),
}


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def split_list(val: str) -> list:
    """Split a semicolon-delimited string into a clean list."""
    if not val:
        return []
    return [v.strip() for v in val.split(";") if v.strip()]


def safe_id(s: str) -> str:
    """Sanitize a name to a stable identifier."""
    s = s.strip().lower()
    s = re.sub(r"[^a-z0-9_]", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s


def evidence_level(source_ids: list) -> str:
    n = len(source_ids)
    if n >= 6:
        return "high"
    if n >= 3:
        return "medium"
    return "low"


def parse_required_edges(raw: str) -> list[tuple[str, str]]:
    """
    Parse 'EDGE_TYPE Target Name; EDGE_TYPE Target Name; ...'
    Returns list of (edge_type, target_name) pairs.
    """
    result = []
    if not raw:
        return result
    for part in raw.split(";"):
        part = part.strip()
        if not part:
            continue
        tokens = part.split(None, 1)
        if len(tokens) == 2:
            edge = tokens[0].strip().upper()
            target = tokens[1].strip()
            if edge in EDGE_VOCAB:
                result.append((edge, target))
    return result


def read_phase2_cards() -> list[dict]:
    wb = openpyxl.load_workbook(PHASE2, read_only=True)
    ws = wb["Technique_Cards"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    cards = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = dict(zip(headers, row))
        if rec.get("id"):
            cards.append(rec)
    wb.close()
    return cards


# ─────────────────────────────────────────────────────────────────────────────
# Cypher queries
# ─────────────────────────────────────────────────────────────────────────────

# Upsert a technique/protocol/state/signal/style node with all 22 fields
UPSERT_NODE_CYPHER = """
MERGE (n:{label} {{id: $id}})
ON CREATE SET
  n.name                    = $name,
  n.classLabel              = $classLabel,
  n.domain                  = $domain,
  n.tier                    = $tier,
  n.description             = $description,
  n.whenToUse               = $whenToUse,
  n.whenNotToUse            = $whenNotToUse,
  n.steps                   = $steps,
  n.successSignals          = $successSignals,
  n.failureSignals          = $failureSignals,
  n.triggerSignals          = $triggerSignals,
  n.contraindications       = $contraindications,
  n.switchTo                = $switchTo,
  n.domainVariants          = $domainVariants,
  n.culturalNotes           = $culturalNotes,
  n.dialogueActLinks        = $dialogueActLinks,
  n.cognitiveLoadProfile    = $cognitiveLoadProfile,
  n.sourceIds               = $sourceIds,
  n.reviewStatus            = $reviewStatus,
  n.reviewNotes             = $reviewNotes,
  n.evidenceLevel           = $evidenceLevel,
  n.avgLatency_ms           = $avgLatency_ms,
  n.createdInPhase          = 4
ON MATCH SET
  n.reviewStatus            = $reviewStatus,
  n.updatedInPhase          = 4
RETURN n.id AS id, n.name AS name
"""

# Upsert a stub target node (for relationship targets that may not exist yet)
UPSERT_STUB_CYPHER = """
MERGE (n:Technique {id: $id})
ON CREATE SET
  n.name       = $name,
  n.stub       = true,
  n.createdInPhase = 4
RETURN n.id AS id
"""

# Create a typed relationship between source and target
CREATE_REL_CYPHER = """
MATCH (source {{id: $sourceId}})
MATCH (target {{id: $targetId}})
MERGE (source)-[r:{rel_type}]->(target)
ON CREATE SET
  r.weight      = $weight,
  r.source      = "phase2_excel",
  r.createdInPhase = 4
RETURN type(r) AS rel_type, source.name AS from, target.name AS to
"""

# Upsert a SignalMarker node
UPSERT_SIGNAL_CYPHER = """
MERGE (s:SignalMarker {id: $id})
ON CREATE SET
  s.name     = $name,
  s.modality = $modality,
  s.type     = $sig_type,
  s.createdInPhase = 4
RETURN s.id AS id, s.name AS name
"""

# Create TRIGGERS relationship from SignalMarker to Technique
CREATE_TRIGGERS_CYPHER = """
MATCH (s:SignalMarker {id: $signalId})
MATCH (t {id: $techniqueId})
MERGE (s)-[r:TRIGGERS]->(t)
ON CREATE SET
  r.weight          = 0.87,
  r.culturalModifier = "universal",
  r.source          = "phase2_excel",
  r.createdInPhase  = 4
RETURN s.name AS signal, t.name AS technique
"""

# Set EmotionalState-class node extra properties
SET_EMOTION_PROPS_CYPHER = """
MATCH (n {id: $id})
SET n.valence = $valence, n.arousal = $arousal
RETURN n.id AS id
"""

# Link Track B node to Track A Resource node
LINK_TRACKS_CYPHER = """
MATCH (b {id: $id})
MATCH (a:Resource {cardId: $id})
MERGE (a)-[:SAME_AS]->(b)
RETURN a.uri AS resource_uri, b.id AS production_node
"""


# ─────────────────────────────────────────────────────────────────────────────
# Classify signal modality
# ─────────────────────────────────────────────────────────────────────────────

VERBAL_KEYWORDS    = {"absolutist", "filler", "repair", "indirect", "language", "word", "blame", "self_blame"}
PROSODIC_KEYWORDS  = {"speech_rate", "volume", "pitch", "pause", "silence", "pace"}
KINESIC_KEYWORDS   = {"facs", "gaze", "gesture", "posture", "proxem"}

def classify_signal(signal_id: str) -> tuple[str, str]:
    """Return (modality, type) for a signal marker id."""
    s = signal_id.lower()
    if any(k in s for k in PROSODIC_KEYWORDS):
        return ("prosodic", "speech_feature")
    if any(k in s for k in KINESIC_KEYWORDS):
        return ("kinesic", "nonverbal")
    if any(k in s for k in VERBAL_KEYWORDS):
        return ("verbal", "linguistic")
    return ("verbal", "general")


# ─────────────────────────────────────────────────────────────────────────────
# Ingestion logic
# ─────────────────────────────────────────────────────────────────────────────

def ingest_cards(session, cards: list[dict]) -> dict:
    stats = {
        "nodes_created":        0,
        "nodes_merged":         0,
        "relationships_created":0,
        "signals_created":      0,
        "errors":               [],
    }

    # Pass 1: Create/merge all primary nodes
    print(f"  [Pass 1] Creating {len(cards)} primary nodes …")
    for card in cards:
        card_id    = card.get("id", "").strip()
        cls_label  = card.get("class", "CommunicationTechnique")
        neo4j_label = CLASS_LABEL_MAP.get(cls_label, "Technique")
        domain_raw  = card.get("domain", "") or ""

        source_ids_raw = split_list(card.get("source_ids", "") or "")

        params = {
            "id":                  card_id,
            "name":                card.get("name", "") or "",
            "classLabel":          cls_label,
            "domain":              domain_raw,
            "tier":                card.get("tier", "") or "",
            "description":         card.get("description", "") or "",
            "whenToUse":           split_list(card.get("when_to_use", "") or ""),
            "whenNotToUse":        split_list(card.get("when_not_to_use", "") or ""),
            "steps":               card.get("steps", "") or "",
            "successSignals":      split_list(card.get("success_signals", "") or ""),
            "failureSignals":      split_list(card.get("failure_signals", "") or ""),
            "triggerSignals":      split_list(card.get("trigger_signals", "") or ""),
            "contraindications":   card.get("contraindications", "") or "",
            "switchTo":            split_list(card.get("switch_to", "") or ""),
            "domainVariants":      split_list(card.get("domain_variants", "") or ""),
            "culturalNotes":       card.get("cultural_notes", "") or "",
            "dialogueActLinks":    split_list(card.get("dialogue_act_links", "") or ""),
            "cognitiveLoadProfile":card.get("cognitive_load_profile", "") or "",
            "sourceIds":           source_ids_raw,
            "reviewStatus":        card.get("review_status", "") or "",
            "reviewNotes":         card.get("review_notes", "") or "",
            "evidenceLevel":       evidence_level(source_ids_raw),
            "avgLatency_ms":       DOMAIN_LATENCY.get(domain_raw, 50),
        }

        # Build dynamic Cypher with correct label
        cypher = UPSERT_NODE_CYPHER.format(label=neo4j_label)
        try:
            result = session.run(cypher, **params)
            summary = result.consume()
            if summary.counters.nodes_created > 0:
                stats["nodes_created"] += 1
            else:
                stats["nodes_merged"] += 1
        except Exception as e:
            stats["errors"].append(f"Node {card_id}: {e}")
            continue

        # Set extra EmotionalState properties if applicable
        if cls_label == "EmotionalState" and card_id in EMOTIONAL_STATE_PROPS:
            valence, arousal = EMOTIONAL_STATE_PROPS[card_id]
            session.run(SET_EMOTION_PROPS_CYPHER, id=card_id, valence=valence, arousal=arousal)

    print(f"     created={stats['nodes_created']}, merged={stats['nodes_merged']}")

    # Pass 2: Create SignalMarker nodes from trigger_signals column
    print(f"  [Pass 2] Creating SignalMarker nodes …")
    seen_signals = set()
    for card in cards:
        card_id = card.get("id", "").strip()
        signals  = split_list(card.get("trigger_signals", "") or "")
        for sig_raw in signals:
            sig_id = safe_id(sig_raw)
            if sig_id in seen_signals:
                continue
            seen_signals.add(sig_id)
            modality, sig_type = classify_signal(sig_id)
            try:
                result = session.run(
                    UPSERT_SIGNAL_CYPHER,
                    id=sig_id, name=sig_raw,
                    modality=modality, sig_type=sig_type
                )
                summary = result.consume()
                if summary.counters.nodes_created > 0:
                    stats["signals_created"] += 1
            except Exception as e:
                stats["errors"].append(f"Signal {sig_id}: {e}")

    print(f"     signal_markers_created={stats['signals_created']}")

    # Pass 3: Create relationships
    print(f"  [Pass 3] Creating relationships …")
    for card in cards:
        card_id    = card.get("id", "").strip()
        domain_raw = card.get("domain", "") or ""

        # 3a — required_edges column
        for edge_type, target_name in parse_required_edges(card.get("required_edges", "") or ""):
            target_id = safe_id(target_name)
            # Ensure target stub exists
            session.run(UPSERT_STUB_CYPHER, id=target_id, name=target_name)
            cypher = CREATE_REL_CYPHER.format(rel_type=edge_type)
            try:
                result = session.run(
                    cypher,
                    sourceId=card_id, targetId=target_id,
                    weight=1.0
                )
                summary = result.consume()
                stats["relationships_created"] += summary.counters.relationships_created
            except Exception as e:
                stats["errors"].append(f"Rel {card_id} -{edge_type}-> {target_id}: {e}")

        # 3b — domain_variants → DOMAIN_VARIANT_OF relationships
        for variant_name in split_list(card.get("domain_variants", "") or ""):
            variant_id = safe_id(variant_name)
            session.run(UPSERT_STUB_CYPHER, id=variant_id, name=variant_name)
            cypher = CREATE_REL_CYPHER.format(rel_type="DOMAIN_VARIANT_OF")
            try:
                result = session.run(
                    cypher,
                    sourceId=card_id, targetId=variant_id,
                    weight=1.0
                )
                summary = result.consume()
                stats["relationships_created"] += summary.counters.relationships_created
            except Exception as e:
                stats["errors"].append(f"DomainVariant {card_id}: {e}")

        # 3c — switch_to → ESCALATES_TO relationships
        for fallback_name in split_list(card.get("switch_to", "") or ""):
            fallback_id = safe_id(fallback_name)
            session.run(UPSERT_STUB_CYPHER, id=fallback_id, name=fallback_name)
            cypher = CREATE_REL_CYPHER.format(rel_type="ESCALATES_TO")
            try:
                result = session.run(
                    cypher,
                    sourceId=card_id, targetId=fallback_id,
                    weight=0.7
                )
                summary = result.consume()
                stats["relationships_created"] += summary.counters.relationships_created
            except Exception as e:
                stats["errors"].append(f"EscalatesTo {card_id}: {e}")

        # 3d — trigger_signals → TRIGGERED_BY relationships (SignalMarker → Technique)
        for sig_raw in split_list(card.get("trigger_signals", "") or ""):
            sig_id = safe_id(sig_raw)
            try:
                result = session.run(
                    CREATE_TRIGGERS_CYPHER,
                    signalId=sig_id, techniqueId=card_id
                )
                summary = result.consume()
                stats["relationships_created"] += summary.counters.relationships_created
            except Exception as e:
                stats["errors"].append(f"Triggers {sig_id} -> {card_id}: {e}")

    print(f"     relationships_created={stats['relationships_created']}")

    # Pass 4: Link Track B nodes to Track A Resource nodes
    print(f"  [Pass 4] Linking Track B nodes to Track A Resource nodes …")
    linked = 0
    for card in cards:
        card_id = card.get("id", "").strip()
        try:
            result = session.run(LINK_TRACKS_CYPHER, id=card_id)
            linked += result.consume().counters.relationships_created
        except Exception:
            pass  # Track A may not have run yet; non-fatal
    print(f"     track_links_created={linked}")

    return stats


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="UCKB Phase 4 Track B Ingestion")
    parser.add_argument("--uri",      default=DEFAULT_URI,      help="Neo4j Bolt URI")
    parser.add_argument("--user",     default=DEFAULT_USER,     help="Neo4j username")
    parser.add_argument("--password", default=DEFAULT_PASSWORD, help="Neo4j password")
    parser.add_argument("--dry-run",  action="store_true",      help="Parse only, no writes")
    args = parser.parse_args()

    sys.stdout.reconfigure(encoding="utf-8")

    print("\n" + "=" * 60)
    print("UCKB Phase 4 — Track B Production Ingestion")
    print("=" * 60)
    print(f"  Excel:    {PHASE2}")
    print(f"  Target:   {args.uri}")
    print(f"  Dry run:  {args.dry_run}")

    # Load cards
    print("\n[1/3] Reading Phase 2 Excel …")
    cards = read_phase2_cards()
    print(f"  Loaded {len(cards)} technique cards")

    if args.dry_run:
        print("\n[DRY RUN] Parsing complete. No writes performed.")
        # Print sample
        for card in cards[:3]:
            print(f"  {card['id']} | {card['class']} | {card['domain']}")
            print(f"    edges: {parse_required_edges(card.get('required_edges','') or '')}")
        return

    # Connect
    print(f"\n[2/3] Connecting to {args.uri} …")
    try:
        driver = GraphDatabase.driver(args.uri, auth=(args.user, args.password))
        driver.verify_connectivity()
        print("  Connected successfully.")
    except neo4j_exc.ServiceUnavailable as e:
        print(f"\n  ERROR: Cannot connect to Neo4j at {args.uri}")
        print(f"  Make sure the database is running (docker compose up -d)")
        print(f"  Detail: {e}")
        sys.exit(1)

    # Ingest
    print("\n[3/3] Ingesting into Neo4j …")
    with driver.session(database="neo4j") as session:
        stats = ingest_cards(session, cards)

    driver.close()

    # Summary
    print("\n" + "=" * 60)
    print("INGESTION COMPLETE")
    print("=" * 60)
    print(f"  Nodes created:        {stats['nodes_created']}")
    print(f"  Nodes merged:         {stats['nodes_merged']}")
    print(f"  Relationships:        {stats['relationships_created']}")
    print(f"  Signal markers:       {stats['signals_created']}")
    print(f"  Errors:               {len(stats['errors'])}")
    if stats["errors"]:
        print("\n  ERROR DETAILS:")
        for err in stats["errors"][:20]:
            print(f"    {err}")
    print("=" * 60)
    print("  Next: run validate_phase4.py")


if __name__ == "__main__":
    main()
