"""
Phase 3 Build Script — UCKB Ontology Design & Semantic Web Technologies
Reads Phase 2 Excel, generates all OWL/Turtle + JSON-LD files, runs SHACL validation.
"""

import os
import re
import sys
import json
import textwrap
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")

import openpyxl
from rdflib import Graph, Namespace, URIRef, Literal, RDF, RDFS, OWL, XSD
from rdflib.namespace import SKOS

# ── Paths ────────────────────────────────────────────────────────────────────
ROOT       = Path(__file__).resolve().parent.parent
PHASE2_XL  = ROOT / "outputs" / "phase2_uckb" / "Phase_2_UCKB_Master_Taxonomy.xlsx"
OUT        = ROOT / "outputs" / "phase3_uckb" / "uckb-ontology"
CORE_DIR   = OUT / "core"
DOM_DIR    = OUT / "domains"
VAL_DIR    = OUT / "validation"
MAP_DIR    = OUT / "mappings"
JLD_DIR    = OUT / "jsonld"
REP_DIR    = OUT / "reports"

for d in [CORE_DIR, DOM_DIR, VAL_DIR, MAP_DIR, JLD_DIR, REP_DIR]:
    d.mkdir(parents=True, exist_ok=True)

# ── Namespaces ───────────────────────────────────────────────────────────────
UCKB = Namespace("https://uckb.io/ontology#")
SAC  = Namespace("https://www.w3.org/community/s-agent-comm/ontology#")
SH   = Namespace("http://www.w3.org/ns/shacl#")
DC   = Namespace("http://purl.org/dc/elements/1.1/")

def base_graph() -> Graph:
    g = Graph()
    g.bind("uckb", UCKB)
    g.bind("owl",  OWL)
    g.bind("rdfs", RDFS)
    g.bind("rdf",  RDF)
    g.bind("xsd",  XSD)
    g.bind("sh",   SH)
    g.bind("sac",  SAC)
    g.bind("skos", SKOS)
    g.bind("dc",   DC)
    return g

def to_iri(s: str) -> str:
    """Sanitize a string to a valid local IRI name."""
    s = s.strip()
    s = re.sub(r"[^a-zA-Z0-9_\-]", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s

def uckb_iri(s: str) -> URIRef:
    return UCKB[to_iri(s)]

def save_ttl(g: Graph, path: Path):
    path.write_text(g.serialize(format="turtle"), encoding="utf-8")
    print(f"  ✓ {path.relative_to(ROOT)}")

def save_jsonld(g: Graph, path: Path):
    path.write_text(g.serialize(format="json-ld", indent=2), encoding="utf-8")

# ─────────────────────────────────────────────────────────────────────────────
# TASK 2  core/communication-acts.ttl
# ─────────────────────────────────────────────────────────────────────────────
def build_communication_acts() -> Graph:
    g = base_graph()

    ont = UCKB["UCKBOntology"]
    g.add((ont, RDF.type, OWL.Ontology))
    g.add((ont, RDFS.label, Literal("Universal Communication Knowledge Base Ontology")))
    g.add((ont, DC["creator"], Literal("UCKB Project")))
    g.add((ont, DC["description"], Literal(
        "Formal OWL ontology encoding all communication techniques, acts, psychological "
        "models, cultural contexts, emotional states, and domain protocols for the UCKB.")))

    # ── 8 top-level OWL classes ──────────────────────────────────────────────
    TOP_CLASSES = {
        "CommunicationAct":
            "The atomic unit — a specific thing said, signaled, or expressed with a defined function.",
        "CommunicationTechnique":
            "A defined, repeatable method for achieving a communication goal.",
        "PsychologicalModel":
            "A theoretical framework explaining human behaviour and emotional state.",
        "CulturalContext":
            "A defined cultural configuration affecting interpretation and norms.",
        "DomainProtocol":
            "A field-specific procedural framework governing communication in a domain.",
        "EmotionalState":
            "A classified affective condition of the interlocutor at a given moment.",
        "CommunicationStyle":
            "A persistent behavioural pattern in how a person communicates.",
        "SignalMarker":
            "An observable verbal, paralinguistic, or behavioural indicator that triggers routing.",
    }
    for cls, desc in TOP_CLASSES.items():
        n = UCKB[cls]
        g.add((n, RDF.type, OWL.Class))
        g.add((n, RDFS.label, Literal(cls)))
        g.add((n, RDFS.comment, Literal(desc)))

    # ── 5 W3C s-agent-comm classes ──────────────────────────────────────────
    SAC_CLASSES = {
        "Agent":           "Any communicative entity (human, AI, organisation) with a defined role and capability set.",
        "Intent":          "The underlying goal or communicative function an agent is attempting to achieve.",
        "Delegation":      "The transfer of a task or authority from one agent to another within a conversation.",
        "Capability":      "The specific communication techniques and knowledge domains an agent has access to.",
        "ExecutionRecord": "A persistent log of what communication act was performed, when, and with what result.",
    }
    for cls, desc in SAC_CLASSES.items():
        n = SAC[cls]
        g.add((n, RDF.type, OWL.Class))
        g.add((n, RDFS.label, Literal(cls)))
        g.add((n, RDFS.comment, Literal(desc)))

    # ── 11 locked object properties (edge types) ────────────────────────────
    EDGE_PROPS = {
        "REQUIRES":             ("Technique requires another technique or state as a prerequisite.", UCKB["CommunicationTechnique"], UCKB["CommunicationTechnique"]),
        "ENHANCES":             ("Applying this technique improves effectiveness of the target.", UCKB["CommunicationTechnique"], UCKB["CommunicationTechnique"]),
        "CONTRADICTS":          ("The two techniques conflict when applied simultaneously.", UCKB["CommunicationTechnique"], UCKB["CommunicationTechnique"]),
        "DOMAIN_VARIANT_OF":    ("This is a domain-specific adaptation of the target core technique.", UCKB["CommunicationTechnique"], UCKB["CommunicationTechnique"]),
        "CULTURAL_VARIANT_OF":  ("This is a culturally adapted form of the target technique.", UCKB["CommunicationTechnique"], UCKB["CommunicationTechnique"]),
        "CONTRAINDICATED_WHEN": ("Technique must not be used when this emotional or contextual state is present.", UCKB["CommunicationTechnique"], UCKB["EmotionalState"]),
        "ESCALATES_TO":         ("Failure of this technique warrants escalation to the target technique.", UCKB["CommunicationTechnique"], UCKB["CommunicationTechnique"]),
        "RESOLVES":             ("This technique resolves or deactivates the target state or dysfunction.", UCKB["CommunicationTechnique"], UCKB["EmotionalState"]),
        "TRIGGERED_BY":         ("This technique is activated when the target signal marker is detected.", UCKB["CommunicationTechnique"], UCKB["SignalMarker"]),
        "PRECEDES":             ("This technique must occur before the target in a protocol sequence.", UCKB["CommunicationTechnique"], UCKB["CommunicationTechnique"]),
        "FOLLOWS":              ("This technique follows the target in a protocol sequence.", UCKB["CommunicationTechnique"], UCKB["CommunicationTechnique"]),
    }
    for prop, (desc, domain, rng) in EDGE_PROPS.items():
        n = UCKB[prop]
        g.add((n, RDF.type, OWL.ObjectProperty))
        g.add((n, RDFS.label, Literal(prop)))
        g.add((n, RDFS.comment, Literal(desc)))
        g.add((n, RDFS.domain, domain))
        g.add((n, RDFS.range, rng))

    # ── 22 data properties (technique card fields) ───────────────────────────
    DATA_PROPS = {
        "cardId":             ("Unique identifier string for this technique card.", XSD.string),
        "name":               ("Human-readable name of the technique.", XSD.string),
        "classLabel":         ("Ontology class this technique belongs to.", XSD.string),
        "domain":             ("Primary domain of application.", XSD.string),
        "tier":               ("Implementation priority tier (1 = highest).", XSD.string),
        "description":        ("Concise definition of what this technique is and what it does.", XSD.string),
        "whenToUse":          ("Conditions under which this technique should be activated.", XSD.string),
        "whenNotToUse":       ("Conditions under which this technique must not be used.", XSD.string),
        "steps":              ("Step-by-step execution protocol.", XSD.string),
        "successSignals":     ("Observable indicators that the technique is working.", XSD.string),
        "failureSignals":     ("Observable indicators that the technique is failing.", XSD.string),
        "triggerSignals":     ("Signal markers whose detection activates this technique.", XSD.string),
        "requiredEdges":      ("Raw edge declarations parsed into object property triples.", XSD.string),
        "contraindications":  ("Safety constraints and absolute restrictions.", XSD.string),
        "switchTo":           ("Fallback techniques if this one fails.", XSD.string),
        "domainVariants":     ("Named domain-specific adaptations of this technique.", XSD.string),
        "culturalNotes":      ("Cultural adaptation notes from Hofstede, Hall, Lewis dimensions.", XSD.string),
        "dialogueActLinks":   ("ISO 24617-2 dialogue act categories this technique maps to.", XSD.string),
        "cognitiveLoadProfile": ("Cognitive load classification (e.g. low-load, high-directive).", XSD.string),
        "sourceIds":          ("Provenance references from the Source Register.", XSD.string),
        "reviewStatus":       ("Current review state of this technique card.", XSD.string),
        "reviewNotes":        ("Free-text review and curation notes.", XSD.string),
    }
    for prop, (desc, dtype) in DATA_PROPS.items():
        n = UCKB[prop]
        g.add((n, RDF.type, OWL.DatatypeProperty))
        g.add((n, RDFS.label, Literal(prop)))
        g.add((n, RDFS.comment, Literal(desc)))
        g.add((n, RDFS.range, dtype))

    return g


# ─────────────────────────────────────────────────────────────────────────────
# TASK 3  core/psychological-models.ttl
# ─────────────────────────────────────────────────────────────────────────────
def build_psychological_models() -> Graph:
    g = base_graph()

    models = [
        ("TransactionalAnalysis",       "Transactional Analysis",
         "Eric Berne's ego-state model (Parent/Adult/Child) governing agent routing via ego-state detection."),
        ("BDI_Model",                   "Belief-Desire-Intention Model",
         "Agent reasoning framework: Belief (current knowledge), Desire (target state), Intention (committed next action)."),
        ("CommunicationAccommodationTheory", "Communication Accommodation Theory",
         "Giles CAT: convergence (rapport-building) and divergence (authority/identity assertion) dynamics."),
        ("RelevanceTheory",             "Relevance Theory",
         "Sperber & Wilson: listeners minimise processing effort while maximising cognitive effects. "
         "Implemented as real-time weight modifier on graph traversal paths."),
        ("KarpmanDramaTriangle",        "Karpman Drama Triangle",
         "Three dysfunctional roles — Victim, Persecutor, Rescuer — that agents must recognise and redirect."),
        ("WinnersTriangle",             "Winner's Triangle",
         "Choy's replacement framework: Vulnerable (from Victim), Assertive (from Persecutor), Caring (from Rescuer)."),
        ("MotivationalInterviewing_Theory", "Motivational Interviewing Theory",
         "Rollnick & Miller: resolving ambivalence by eliciting change talk and rolling with resistance."),
        ("AttachmentTheory",            "Attachment Theory",
         "Bowlby/Ainsworth: secure, anxious, avoidant, and disorganised attachment styles affecting rapport strategies."),
        ("CognitiveLoadTheory",         "Cognitive Load Theory",
         "Sweller: intrinsic, extraneous, germane load — governs Relevance Theory routing in UCKB."),
        ("SocialPenetrationTheory",     "Social Penetration Theory",
         "Altman & Taylor: self-disclosure depth and breadth as trust-building signals."),
        ("PolyVagalTheory",             "Polyvagal Theory",
         "Porges: autonomic nervous system states (ventral vagal, sympathetic, dorsal vagal) mapped to de-escalation protocols."),
    ]

    for iri, label, comment in models:
        n = UCKB[iri]
        g.add((n, RDF.type, OWL.NamedIndividual))
        g.add((n, RDF.type, UCKB["PsychologicalModel"]))
        g.add((n, RDFS.label, Literal(label)))
        g.add((n, RDFS.comment, Literal(comment)))

    return g


# ─────────────────────────────────────────────────────────────────────────────
# TASK 4  core/cultural-contexts.ttl
# ─────────────────────────────────────────────────────────────────────────────
def build_cultural_contexts() -> Graph:
    g = base_graph()

    contexts = [
        # Hofstede dimensions
        ("PowerDistanceHigh",         "High Power Distance",
         "Hierarchical authority structures accepted; formal communication preferred."),
        ("PowerDistanceLow",          "Low Power Distance",
         "Flat authority structures; informal, egalitarian communication expected."),
        ("IndividualismHigh",         "High Individualism",
         "Personal goals, autonomy, and direct I-statements dominate communication."),
        ("CollectivismHigh",          "High Collectivism",
         "In-group harmony, face-saving, and indirect communication preferred."),
        ("UncertaintyAvoidanceHigh",  "High Uncertainty Avoidance",
         "Preference for rules, formality, and explicit reassurance."),
        ("UncertaintyAvoidanceLow",   "Low Uncertainty Avoidance",
         "Tolerance for ambiguity; flexible, improvisational style accepted."),
        ("LongTermOrientation",       "Long-Term Orientation",
         "Perseverance and relationship investment valued over immediate results."),
        ("ShortTermOrientation",      "Short-Term Orientation",
         "Immediate results and face maintenance prioritised."),
        ("MasculinityHigh",           "High Masculinity",
         "Achievement, assertiveness, and material success dominate communication goals."),
        ("FemininityHigh",            "High Femininity",
         "Cooperation, modesty, caring, and quality of life shape communication register."),
        # Hall dimensions
        ("HighContext",               "High Context",
         "Meaning conveyed through implicit cues, relationship history, and nonverbal signals."),
        ("LowContext",                "Low Context",
         "Meaning conveyed explicitly through words; directness expected."),
        ("MonochronicTime",           "Monochronic Time",
         "Linear, schedule-driven; interruption is disrespectful."),
        ("PolychronicTime",           "Polychronic Time",
         "Parallel activities and relationship flow over rigid schedules."),
        ("HighProxemics",             "High Proxemics",
         "Close physical and conversational space is normal; distance signals rejection."),
        ("LowProxemics",              "Low Proxemics",
         "Larger personal space norms; proximity may feel intrusive."),
        # Lewis model
        ("LinearActive",              "Linear-Active",
         "Task-oriented, logical, direct; expects factual, structured communication."),
        ("MultiActive",               "Multi-Active",
         "Relationship-oriented, emotional, spontaneous; rapport before task."),
        ("Reactive",                  "Reactive",
         "Listens carefully before responding; values harmony, indirect communication, face-saving."),
    ]

    for iri, label, comment in contexts:
        n = UCKB[iri]
        g.add((n, RDF.type, OWL.NamedIndividual))
        g.add((n, RDF.type, UCKB["CulturalContext"]))
        g.add((n, RDFS.label, Literal(label)))
        g.add((n, RDFS.comment, Literal(comment)))

    return g


# ─────────────────────────────────────────────────────────────────────────────
# TASK 5  core/emotional-states.ttl
# ─────────────────────────────────────────────────────────────────────────────
def build_emotional_states() -> Graph:
    g = base_graph()

    # (iri, label, valence, arousal, comment)
    states = [
        ("Panic",              "Panic",               "negative", "high",
         "Acute fear response with cognitive fragmentation and fight-or-flight activation."),
        ("Grief",              "Grief",               "negative", "low",
         "Profound sadness following loss; requires validation before information exchange."),
        ("Anger",              "Anger",               "negative", "high",
         "High-arousal negative affect; may present as volume spike or absolutist language."),
        ("Hostile",            "Hostile",             "negative", "high",
         "Sustained antagonistic orientation toward the agent."),
        ("Defensive",          "Defensive",           "negative", "medium",
         "Protective stance against perceived criticism or threat."),
        ("Dissociated",        "Dissociated",         "negative", "low",
         "Disconnection from emotional or physical experience; flat affect, minimal response."),
        ("Ambivalent",         "Ambivalent",          "neutral",  "medium",
         "Co-existing motivations for and against change; target state for MI techniques."),
        ("Resistant",          "Resistant",           "negative", "medium",
         "Active opposition to agent suggestions or protocol steps."),
        ("Compliant",          "Compliant",           "positive", "low",
         "Cooperative engagement with agent requests."),
        ("Receptive",          "Receptive",           "positive", "medium",
         "Open and ready to receive information or guidance."),
        ("Suicidal_Ideation",  "Suicidal Ideation",   "negative", "high",
         "Active suicidal ideation; requires immediate safety protocol activation."),
        ("Contemptuous",       "Contemptuous",        "negative", "medium",
         "Expressed contempt — a strong predictor of relationship or compliance breakdown."),
        ("Shame",              "Shame",               "negative", "low",
         "Inward-directed self-attack; may silence or immobilise the interlocutor."),
        ("Fear",               "Fear",                "negative", "high",
         "Threat-activated state; reduces cognitive bandwidth."),
        ("Hopeless",           "Hopeless",            "negative", "low",
         "Belief that change is impossible; may require Motivational Interviewing."),
        ("Confused",           "Confused",            "neutral",  "medium",
         "Disorientation and cognitive overload; triggers simplification routing."),
        ("Overwhelmed",        "Overwhelmed",         "negative", "high",
         "Cognitive and emotional resources exhausted; grounding required."),
        ("CognitiveOverload",  "Cognitive Overload",  "negative", "high",
         "Exceeds working memory capacity; requires reduction in linguistic density."),
        ("ActivePsychoticEpisode", "Active Psychotic Episode", "negative", "high",
         "Contraindication for persuasion and validation techniques requiring insight."),
        ("Calm",               "Calm",                "positive", "low",
         "Regulated baseline state; full cognitive bandwidth available."),
        ("Distress",           "Distress",            "negative", "high",
         "General high-arousal negative affect not yet classified to specific emotion."),
        ("Helpless",           "Helpless",            "negative", "low",
         "Perceived loss of agency; Victim-role marker in Karpman Drama Triangle."),
        ("Denial",             "Denial",              "negative", "medium",
         "Refusal to acknowledge facts or emotional reality."),
        ("Guilt",              "Guilt",               "negative", "medium",
         "Other-directed sense of wrongdoing; different routing from Shame."),
        ("Vulnerable",         "Vulnerable",          "neutral",  "low",
         "Open and emotionally exposed; requires protective framing."),
    ]

    for iri, label, valence, arousal, comment in states:
        n = UCKB[iri]
        g.add((n, RDF.type, OWL.NamedIndividual))
        g.add((n, RDF.type, UCKB["EmotionalState"]))
        g.add((n, RDFS.label, Literal(label)))
        g.add((n, RDFS.comment, Literal(comment)))
        g.add((n, UCKB["valence"], Literal(valence, datatype=XSD.string)))
        g.add((n, UCKB["arousal"], Literal(arousal, datatype=XSD.string)))

    return g


# ─────────────────────────────────────────────────────────────────────────────
# TASK 6  core/communication-styles.ttl
# ─────────────────────────────────────────────────────────────────────────────
def build_communication_styles() -> Graph:
    g = base_graph()

    styles = [
        ("Assertive",        "Assertive",
         "Direct, confident, and respectful; expresses needs and boundaries clearly."),
        ("Passive",          "Passive",
         "Avoids conflict; own needs minimised; may build resentment over time."),
        ("Aggressive",       "Aggressive",
         "Dominates, interrupts, and disregards others' boundaries."),
        ("PassiveAggressive","Passive-Aggressive",
         "Indirect hostility; superficial compliance with covert resistance."),
        ("Analytical",       "Analytical",
         "Data-driven, systematic, detail-oriented; requires evidence before commitment."),
        ("Expressive",       "Expressive",
         "Emotional, spontaneous, relationship-focused; high interpersonal energy."),
        ("Driver",           "Driver",
         "Results-oriented, decisive, direct; low tolerance for ambiguity or delay."),
        ("Amiable",          "Amiable",
         "Cooperative, empathic, consensus-seeking; avoids confrontation."),
        ("Formal",           "Formal",
         "Protocol-bound, structured, distance-maintaining."),
        ("Informal",         "Informal",
         "Casual, rapport-centred, flexible register."),
    ]

    for iri, label, comment in styles:
        n = UCKB[iri]
        g.add((n, RDF.type, OWL.NamedIndividual))
        g.add((n, RDF.type, UCKB["CommunicationStyle"]))
        g.add((n, RDFS.label, Literal(label)))
        g.add((n, RDFS.comment, Literal(comment)))

    return g


# ─────────────────────────────────────────────────────────────────────────────
# TASK 7-9  domains/*.ttl  — build from Phase 2 Excel
# ─────────────────────────────────────────────────────────────────────────────

DOMAIN_MAP = {
    "Crisis Dispatch / Emergency": "dispatch",
    "Sales & Negotiation":         "negotiation",
    "Clinical / Medical":          "clinical",
}

EDGE_VOCAB = {
    "REQUIRES", "ENHANCES", "CONTRADICTS", "DOMAIN_VARIANT_OF",
    "CULTURAL_VARIANT_OF", "CONTRAINDICATED_WHEN", "ESCALATES_TO",
    "RESOLVES", "TRIGGERED_BY", "PRECEDES", "FOLLOWS",
}

# valence/arousal for EmotionalState-class technique-card individuals
EMOTIONAL_STATE_PROPS = {
    "crisis_dispatch_026_panic_state":          ("negative", "high"),
    "crisis_dispatch_027_hostile_state":         ("negative", "high"),
    "crisis_dispatch_028_defensive_state":       ("negative", "medium"),
    "crisis_dispatch_029_cognitive_overload_state": ("negative", "high"),
    "crisis_dispatch_048_adult_ego_state":       ("neutral",  "medium"),
    "clinical_032_grief_pause":                  ("negative", "low"),
    "clinical_033_anger_at_diagnosis_validation":("negative", "high"),
    "clinical_034_fear_of_prognosis_validation": ("negative", "high"),
    "clinical_045_clinical_panic_state":         ("negative", "high"),
}

def parse_edges(raw: str):
    """Parse 'EDGE_TYPE Target Name; ...' into list of (edge, target_iri) tuples."""
    triples = []
    if not raw:
        return triples
    for part in raw.split(";"):
        part = part.strip()
        if not part:
            continue
        tokens = part.split(None, 1)
        if len(tokens) == 2:
            edge, target = tokens[0].strip().upper(), tokens[1].strip()
            if edge in EDGE_VOCAB:
                triples.append((UCKB[edge], uckb_iri(target)))
    return triples

def parse_variants(raw: str):
    """Parse semicolon-separated variant names into list of IRIs."""
    if not raw:
        return []
    return [uckb_iri(v.strip()) for v in raw.split(";") if v.strip()]

def build_domain_graphs():
    wb = openpyxl.load_workbook(PHASE2_XL, read_only=True)
    ws = wb["Technique_Cards"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

    graphs = {slug: base_graph() for slug in DOMAIN_MAP.values()}
    counts = {slug: 0 for slug in DOMAIN_MAP.values()}

    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = dict(zip(headers, row))
        domain_raw = rec.get("domain", "") or ""
        slug = DOMAIN_MAP.get(domain_raw.strip())
        if slug is None:
            continue

        g = graphs[slug]
        card_id = rec.get("id", "") or ""
        n = UCKB[to_iri(card_id)]

        # rdf:type
        g.add((n, RDF.type, OWL.NamedIndividual))
        cls_label = rec.get("class", "CommunicationTechnique") or "CommunicationTechnique"
        g.add((n, RDF.type, UCKB[cls_label]))

        # rdfs:label
        g.add((n, RDFS.label, Literal(str(rec.get("name", "") or ""))))

        # All 22 data properties
        field_map = [
            ("id",                   "cardId"),
            ("name",                 "name"),
            ("class",                "classLabel"),
            ("domain",               "domain"),
            ("tier",                 "tier"),
            ("description",          "description"),
            ("when_to_use",          "whenToUse"),
            ("when_not_to_use",      "whenNotToUse"),
            ("steps",                "steps"),
            ("success_signals",      "successSignals"),
            ("failure_signals",      "failureSignals"),
            ("trigger_signals",      "triggerSignals"),
            ("required_edges",       "requiredEdges"),
            ("contraindications",    "contraindications"),
            ("switch_to",            "switchTo"),
            ("domain_variants",      "domainVariants"),
            ("cultural_notes",       "culturalNotes"),
            ("dialogue_act_links",   "dialogueActLinks"),
            ("cognitive_load_profile","cognitiveLoadProfile"),
            ("source_ids",           "sourceIds"),
            ("review_status",        "reviewStatus"),
            ("review_notes",         "reviewNotes"),
        ]
        for xl_col, owl_prop in field_map:
            val = rec.get(xl_col)
            if val is not None and str(val).strip():
                g.add((n, UCKB[owl_prop], Literal(str(val).strip(), datatype=XSD.string)))

        # Object property triples from required_edges
        for edge_prop, target_iri in parse_edges(rec.get("required_edges", "") or ""):
            g.add((n, edge_prop, target_iri))

        # DOMAIN_VARIANT_OF triples from domain_variants
        for variant_iri in parse_variants(rec.get("domain_variants", "") or ""):
            g.add((n, UCKB["DOMAIN_VARIANT_OF"], variant_iri))

        # valence/arousal for EmotionalState-class individuals
        if cls_label == "EmotionalState" and card_id in EMOTIONAL_STATE_PROPS:
            valence, arousal = EMOTIONAL_STATE_PROPS[card_id]
            g.add((n, UCKB["valence"], Literal(valence, datatype=XSD.string)))
            g.add((n, UCKB["arousal"], Literal(arousal, datatype=XSD.string)))

        counts[slug] += 1

    wb.close()
    return graphs, counts


# ─────────────────────────────────────────────────────────────────────────────
# TASK 10  validation/shacl-shapes.ttl
# ─────────────────────────────────────────────────────────────────────────────
def build_shacl_shapes() -> Graph:
    g = base_graph()

    shape = UCKB["TechniqueShape"]
    g.add((shape, RDF.type, SH.NodeShape))
    g.add((shape, SH.targetClass, UCKB["CommunicationTechnique"]))
    g.add((shape, RDFS.label, Literal("UCKB CommunicationTechnique mandatory property shape")))

    def add_property_shape(path, min_count, max_count, severity, message, in_list=None):
        ps = UCKB[f"PS_{to_iri(str(path).split('#')[-1])}_{min_count}"]
        g.add((shape, SH.property, ps))
        g.add((ps, RDF.type, SH.PropertyShape))
        g.add((ps, SH.path, path))
        g.add((ps, SH.minCount, Literal(min_count, datatype=XSD.integer)))
        if max_count is not None:
            g.add((ps, SH.maxCount, Literal(max_count, datatype=XSD.integer)))
        g.add((ps, SH.severity, severity))
        g.add((ps, SH.message, Literal(message)))
        if in_list:
            list_node = g.skolemize()  # blank node workaround
            from rdflib.collection import Collection
            Collection(g, list_node, [Literal(v) for v in in_list])
            g.add((ps, SH["in"], list_node))
        return ps

    # Shape 1: must have rdfs:label
    add_property_shape(RDFS.label, 1, 1, SH.Violation,
        "Every CommunicationTechnique must have exactly one rdfs:label.")

    # Shape 2: whenToUse mandatory
    add_property_shape(UCKB["whenToUse"], 1, None, SH.Violation,
        "whenToUse is mandatory on every CommunicationTechnique.")

    # Shape 3: whenNotToUse mandatory
    add_property_shape(UCKB["whenNotToUse"], 1, None, SH.Violation,
        "whenNotToUse is mandatory on every CommunicationTechnique.")

    # Shape 4: domain mandatory
    add_property_shape(UCKB["domain"], 1, None, SH.Violation,
        "domain must be specified for every CommunicationTechnique.")

    # Shape 5: triggerSignals — at least one (warning if missing)
    add_property_shape(UCKB["triggerSignals"], 1, None, SH.Warning,
        "CommunicationTechnique should have at least one triggerSignals entry.")

    # Shape 6: steps mandatory
    add_property_shape(UCKB["steps"], 1, None, SH.Violation,
        "steps is mandatory on every CommunicationTechnique.")

    # Shape 7: cardId unique identifier
    add_property_shape(UCKB["cardId"], 1, 1, SH.Violation,
        "Every CommunicationTechnique must have exactly one cardId.")

    # Shape 8: cognitiveLoadProfile mandatory
    add_property_shape(UCKB["cognitiveLoadProfile"], 1, None, SH.Violation,
        "cognitiveLoadProfile is mandatory on every CommunicationTechnique.")

    # Shape 9: sourceIds mandatory
    add_property_shape(UCKB["sourceIds"], 1, None, SH.Violation,
        "sourceIds is mandatory — every technique must have provenance.")

    # ── DomainProtocol shape ─────────────────────────────────────────────────
    dp_shape = UCKB["DomainProtocolShape"]
    g.add((dp_shape, RDF.type, SH.NodeShape))
    g.add((dp_shape, SH.targetClass, UCKB["DomainProtocol"]))
    ps_dp = UCKB["PS_DomainProtocol_label"]
    g.add((dp_shape, SH.property, ps_dp))
    g.add((ps_dp, RDF.type, SH.PropertyShape))
    g.add((ps_dp, SH.path, RDFS.label))
    g.add((ps_dp, SH.minCount, Literal(1, datatype=XSD.integer)))
    g.add((ps_dp, SH.severity, SH.Violation))
    g.add((ps_dp, SH.message, Literal("Every DomainProtocol must have rdfs:label.")))

    # ── EmotionalState shape ─────────────────────────────────────────────────
    es_shape = UCKB["EmotionalStateShape"]
    g.add((es_shape, RDF.type, SH.NodeShape))
    g.add((es_shape, SH.targetClass, UCKB["EmotionalState"]))
    for prop_iri, msg in [
        (RDFS.label,        "EmotionalState must have rdfs:label."),
        (UCKB["valence"],   "EmotionalState must have valence (positive/negative/neutral)."),
        (UCKB["arousal"],   "EmotionalState must have arousal (high/medium/low)."),
    ]:
        ps = UCKB[f"PS_ES_{to_iri(str(prop_iri).split('#')[-1])}"]
        g.add((es_shape, SH.property, ps))
        g.add((ps, RDF.type, SH.PropertyShape))
        g.add((ps, SH.path, prop_iri))
        g.add((ps, SH.minCount, Literal(1, datatype=XSD.integer)))
        g.add((ps, SH.severity, SH.Violation))
        g.add((ps, SH.message, Literal(msg)))

    return g


# ─────────────────────────────────────────────────────────────────────────────
# TASK 11  mappings/domain-to-core-mappings.ttl
# ─────────────────────────────────────────────────────────────────────────────
def build_mappings() -> Graph:
    g = base_graph()

    # Core technique anchors
    core_anchors = [
        ("ActiveListening",     "Active Listening",     "CommunicationTechnique",
         "Core universal active listening framework."),
        ("EmpathicValidation",  "Empathic Validation",  "CommunicationTechnique",
         "Core universal empathic validation framework."),
        ("Grounding",           "Grounding",            "CommunicationTechnique",
         "Core present-focus grounding framework."),
        ("Rapport",             "Rapport",              "CommunicationTechnique",
         "Core rapport-building framework."),
        ("Empathy",             "Empathy",              "CommunicationTechnique",
         "Core empathy expression framework."),
        ("MotivationalInterviewing", "Motivational Interviewing", "DomainProtocol",
         "Core MI protocol for ambivalence resolution."),
        ("NeedsAssessment",     "Needs Assessment",     "CommunicationTechnique",
         "Core needs-elicitation framework."),
        ("Negotiation",         "Negotiation",          "DomainProtocol",
         "Core negotiation framework."),
        ("InformationGathering","Information Gathering", "CommunicationTechnique",
         "Core structured elicitation framework."),
        ("ConflictDeEscalation","Conflict De-escalation","CommunicationTechnique",
         "Core de-escalation framework."),
    ]
    for iri, label, cls, comment in core_anchors:
        n = UCKB[iri]
        g.add((n, RDF.type, OWL.NamedIndividual))
        g.add((n, RDF.type, UCKB[cls]))
        g.add((n, RDFS.label, Literal(label)))
        g.add((n, RDFS.comment, Literal(comment)))

    # Domain-to-core variant mappings (from Phase 2 domain_variants column analysis)
    variant_links = [
        # Crisis dispatch variants → core
        ("TacticalEmpathy",          "ActiveListening"),
        ("SOLER",                    "ActiveListening"),
        ("ReflectiveListening",      "ActiveListening"),
        ("CrisisEmpathy",            "EmpathicValidation"),
        ("TraumaInformedGrounding",  "Grounding"),
        ("ClinicalGrounding",        "Grounding"),
        # Sales & Negotiation variants → core
        ("SPIN_Selling",             "NeedsAssessment"),
        ("ChallengerInsight",        "NeedsAssessment"),
        ("HarvardPrincipledNeg",     "Negotiation"),
        ("SandlerPainFunnel",        "NeedsAssessment"),
        ("MEDDIC_Qualification",     "NeedsAssessment"),
        ("NegotiationRapport",       "Rapport"),
        ("SalesEmpathy",             "EmpathicValidation"),
        # Clinical variants → core
        ("SPIKES_SupportStep",       "EmpathicValidation"),
        ("MotivationalListening",    "ActiveListening"),
        ("TeachBackVerification",    "InformationGathering"),
        ("ClinicalEmpathy",          "EmpathicValidation"),
        ("SafetyPlanning",           "ConflictDeEscalation"),
        ("CollaborativeProblSolving","NeedsAssessment"),
    ]
    for variant, core in variant_links:
        v_iri = UCKB[variant]
        c_iri = UCKB[core]
        g.add((v_iri, RDF.type, OWL.NamedIndividual))
        g.add((v_iri, UCKB["DOMAIN_VARIANT_OF"], c_iri))

    return g


# ─────────────────────────────────────────────────────────────────────────────
# JSON-LD export
# ─────────────────────────────────────────────────────────────────────────────
def export_jsonld(source_ttl: Path):
    g = Graph()
    g.parse(str(source_ttl), format="turtle")
    stem = source_ttl.stem
    out_path = JLD_DIR / f"{stem}.jsonld"
    save_jsonld(g, out_path)
    print(f"  ✓ jsonld/{stem}.jsonld")


# ─────────────────────────────────────────────────────────────────────────────
# SHACL validation
# ─────────────────────────────────────────────────────────────────────────────
def run_shacl_validation(domain_ttl_paths, shapes_path):
    import pyshacl

    report_lines = ["UCKB Phase 3 — SHACL Validation Report", "=" * 60, ""]
    total_violations = 0
    total_warnings   = 0
    all_pass = True

    for ttl_path in domain_ttl_paths:
        data_graph = Graph()
        data_graph.parse(str(ttl_path), format="turtle")

        shapes_graph = Graph()
        shapes_graph.parse(str(shapes_path), format="turtle")

        conforms, results_graph, results_text = pyshacl.validate(
            data_graph,
            shacl_graph=shapes_graph,
            inference="rdfs",
            abort_on_first=False,
            allow_infos=True,
            meta_shacl=False,
            debug=False,
        )

        viol = results_text.count("Constraint Violation")
        warn = results_text.count("Constraint Warning")
        total_violations += viol
        total_warnings   += warn

        status = "PASS" if conforms else "FAIL"
        if not conforms:
            all_pass = False

        report_lines.append(f"File : {ttl_path.name}")
        report_lines.append(f"Status   : {status}")
        report_lines.append(f"Violations: {viol}  |  Warnings: {warn}")
        report_lines.append("")
        if not conforms:
            report_lines.append("Detail:")
            report_lines.append(results_text)
            report_lines.append("")

    report_lines.append("=" * 60)
    report_lines.append(f"TOTAL VIOLATIONS : {total_violations}")
    report_lines.append(f"TOTAL WARNINGS   : {total_warnings}")
    report_lines.append(f"OVERALL          : {'ALL PASS' if all_pass else 'FAILURES PRESENT'}")

    report_text = "\n".join(report_lines)
    report_file = REP_DIR / "shacl_validation_report.txt"
    report_file.write_text(report_text, encoding="utf-8")
    print(f"\n  SHACL: violations={total_violations}, warnings={total_warnings}")
    print(f"  Report → {report_file.relative_to(ROOT)}")
    return all_pass, total_violations, total_warnings


# ─────────────────────────────────────────────────────────────────────────────
# Acceptance criteria audit
# ─────────────────────────────────────────────────────────────────────────────
def run_acceptance_audit(domain_counts, shacl_pass, violations, warnings):
    print("\n" + "=" * 60)
    print("PHASE 3 ACCEPTANCE CRITERIA")
    print("=" * 60)

    total_cards = sum(domain_counts.values())
    checks = [
        ("AC-1", "150 technique cards across domain files",
         total_cards >= 150, f"{total_cards}/150"),
        ("AC-2", "Crisis dispatch: 60 cards",
         domain_counts.get("dispatch", 0) == 60, str(domain_counts.get("dispatch", 0))),
        ("AC-3", "Sales & Negotiation: 45 cards",
         domain_counts.get("negotiation", 0) == 45, str(domain_counts.get("negotiation", 0))),
        ("AC-4", "Clinical / Medical: 45 cards",
         domain_counts.get("clinical", 0) == 45, str(domain_counts.get("clinical", 0))),
        ("AC-5", "SHACL shapes: zero violations",
         violations == 0, f"{violations} violations"),
        ("AC-6", "All .ttl files serialised",
         all((CORE_DIR / f).exists() for f in [
             "communication-acts.ttl", "psychological-models.ttl",
             "cultural-contexts.ttl", "emotional-states.ttl",
             "communication-styles.ttl"
         ]), "core/ files"),
        ("AC-7", "All domain .ttl files serialised",
         all((DOM_DIR / f).exists() for f in
             ["dispatch.ttl", "negotiation.ttl", "clinical.ttl"]),
         "domains/ files"),
        ("AC-8", "SHACL shapes file present",
         (VAL_DIR / "shacl-shapes.ttl").exists(), "validation/shacl-shapes.ttl"),
        ("AC-9", "Mappings file present",
         (MAP_DIR / "domain-to-core-mappings.ttl").exists(), "mappings/ file"),
        ("AC-10","All JSON-LD files generated",
         len(list(JLD_DIR.glob("*.jsonld"))) >= 8, f"{len(list(JLD_DIR.glob('*.jsonld')))} files"),
    ]

    all_pass = True
    for ac_id, description, result, detail in checks:
        status = "PASS" if result else "FAIL"
        if not result:
            all_pass = False
        print(f"  {ac_id}  [{status}]  {description}  ({detail})")

    print("=" * 60)
    print(f"  OVERALL: {'ALL PASS ✓' if all_pass else 'FAILURES PRESENT ✗'}")
    print("=" * 60)
    return all_pass


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────
def main():
    print("\n" + "=" * 60)
    print("PHASE 3  UCKB Ontology Build")
    print("=" * 60)

    print("\n[1/5] Building core ontology files …")
    save_ttl(build_communication_acts(),  CORE_DIR / "communication-acts.ttl")
    save_ttl(build_psychological_models(), CORE_DIR / "psychological-models.ttl")
    save_ttl(build_cultural_contexts(),   CORE_DIR / "cultural-contexts.ttl")
    save_ttl(build_emotional_states(),    CORE_DIR / "emotional-states.ttl")
    save_ttl(build_communication_styles(),CORE_DIR / "communication-styles.ttl")

    print("\n[2/5] Building domain ontology files from Phase 2 Excel …")
    domain_graphs, counts = build_domain_graphs()
    save_ttl(domain_graphs["dispatch"],    DOM_DIR / "dispatch.ttl")
    save_ttl(domain_graphs["negotiation"], DOM_DIR / "negotiation.ttl")
    save_ttl(domain_graphs["clinical"],    DOM_DIR / "clinical.ttl")
    for slug, n in counts.items():
        print(f"       {slug}: {n} technique individuals")

    print("\n[3/5] Building SHACL validation shapes …")
    save_ttl(build_shacl_shapes(), VAL_DIR / "shacl-shapes.ttl")

    print("\n[4/5] Building domain-to-core mappings …")
    save_ttl(build_mappings(), MAP_DIR / "domain-to-core-mappings.ttl")

    print("\n[5/5] Exporting JSON-LD …")
    for ttl_path in sorted(list(CORE_DIR.glob("*.ttl")) +
                           list(DOM_DIR.glob("*.ttl")) +
                           list(VAL_DIR.glob("*.ttl")) +
                           list(MAP_DIR.glob("*.ttl"))):
        export_jsonld(ttl_path)

    print("\n[6/6] Running SHACL validation …")
    domain_ttls = [DOM_DIR / "dispatch.ttl",
                   DOM_DIR / "negotiation.ttl",
                   DOM_DIR / "clinical.ttl"]
    shacl_pass, violations, warnings = run_shacl_validation(
        domain_ttls, VAL_DIR / "shacl-shapes.ttl"
    )

    run_acceptance_audit(counts, shacl_pass, violations, warnings)


if __name__ == "__main__":
    main()
